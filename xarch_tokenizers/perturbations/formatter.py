## Written by Claude
import random
from enum import Enum
from typing import Any, Dict, List, Optional

import numpy as np


# Assuming LANGS is defined elsewhere
class LANGS(Enum):
    eng_Latn = "eng_Latn"
    tur_Latn = "tur_Latn"
    zho_Hans = "zho_Hans"
    fas_Arab = "fas_Arab"
    ita_Latn = "ita_Latn"


# =============================================================================
# UNICODE FORMATTING SYSTEM
# =============================================================================


class UNICODE_FORMAT_STYLE(Enum):
    STRIKETHROUGH = "strikethrough"  # S̷t̷r̷i̷k̷e̷
    UNDERLINE = "underline"  # U̲n̲d̲e̲r̲l̲i̲n̲e̲
    SUPERSCRIPT = "superscript"  # ˢᵘᵖᵉʳˢᶜʳⁱᵖᵗ
    SUBSCRIPT = "subscript"  # ₛᵤᵦₛcᵣᵢₚₜ
    UPSIDE_DOWN = "upside_down"  # uʍop ǝpᴉsdn
    WIDE_SPACING = "wide_spacing"  # w i d e   s p a c i n g
    BLOCK_LETTERS = "block"  # █B█l█o█c█k█
    INVISIBLE_SEPARATOR = "invisible"  # i‍n‍v‍i‍s‍i‍b‍l‍e
    ZERO_WIDTH = "zero_width"  # z‎e‎r‎o‎ ‎w‎i‎d‎t‎h
    HAIR_SPACE = "hair_space"  # h‏a‏i‏r‏ ‏s‏p‏a‏c‏e
    THIN_SPACE = "thin_space"  # t h i n   s p a c e
    MIXED_CASE = "mixed_case"  # MiXeD CaSe


class UnicodeFormatter:
    def __init__(self):
        # Unicode combining characters
        self.combining_chars = {
            UNICODE_FORMAT_STYLE.STRIKETHROUGH: "\u0337",  # ̷
            UNICODE_FORMAT_STYLE.UNDERLINE: "\u0332",  # ̲
        }

        # Character mappings
        self.superscript_map = {
            "a": "ᵃ",
            "b": "ᵇ",
            "c": "ᶜ",
            "d": "ᵈ",
            "e": "ᵉ",
            "f": "ᶠ",
            "g": "ᵍ",
            "h": "ʰ",
            "i": "ⁱ",
            "j": "ʲ",
            "k": "ᵏ",
            "l": "ˡ",
            "m": "ᵐ",
            "n": "ⁿ",
            "o": "ᵒ",
            "p": "ᵖ",
            "q": "q",
            "r": "ʳ",
            "s": "ˢ",
            "t": "ᵗ",
            "u": "ᵘ",
            "v": "ᵛ",
            "w": "ʷ",
            "x": "ˣ",
            "y": "ʸ",
            "z": "ᶻ",
            "A": "ᴬ",
            "B": "ᴮ",
            "C": "ᶜ",
            "D": "ᴰ",
            "E": "ᴱ",
            "F": "ᶠ",
            "G": "ᴳ",
            "H": "ᴴ",
            "I": "ᴵ",
            "J": "ᴶ",
            "K": "ᴷ",
            "L": "ᴸ",
            "M": "ᴹ",
            "N": "ᴺ",
            "O": "ᴼ",
            "P": "ᴾ",
            "Q": "Q",
            "R": "ᴿ",
            "S": "ˢ",
            "T": "ᵀ",
            "U": "ᵁ",
            "V": "ⱽ",
            "W": "ᵂ",
            "X": "ˣ",
            "Y": "ʸ",
            "Z": "ᶻ",
        }

        self.subscript_map = {
            "a": "ₐ",
            "e": "ₑ",
            "h": "ₕ",
            "i": "ᵢ",
            "j": "ⱼ",
            "k": "ₖ",
            "l": "ₗ",
            "m": "ₘ",
            "n": "ₙ",
            "o": "ₒ",
            "p": "ₚ",
            "r": "ᵣ",
            "s": "ₛ",
            "t": "ₜ",
            "u": "ᵤ",
            "v": "ᵥ",
            "x": "ₓ",
            "0": "₀",
            "1": "₁",
            "2": "₂",
            "3": "₃",
            "4": "₄",
            "5": "₅",
            "6": "₆",
            "7": "₇",
            "8": "₈",
            "9": "₉",
            "+": "₊",
            "-": "₋",
            "=": "₌",
            "(": "₍",
            ")": "₎",
        }

        self.upside_down_map = {
            "a": "ɐ",
            "b": "q",
            "c": "ɔ",
            "d": "p",
            "e": "ǝ",
            "f": "ɟ",
            "g": "ƃ",
            "h": "ɥ",
            "i": "ᴉ",
            "j": "ſ",
            "k": "ʞ",
            "l": "l",
            "m": "ɯ",
            "n": "u",
            "o": "o",
            "p": "d",
            "q": "b",
            "r": "ɹ",
            "s": "s",
            "t": "ʇ",
            "u": "n",
            "v": "ʌ",
            "w": "ʍ",
            "x": "x",
            "y": "ʎ",
            "z": "z",
            "A": "∀",
            "B": "ᗺ",
            "C": "Ɔ",
            "D": "ᗡ",
            "E": "Ǝ",
            "F": "ᖴ",
            "G": "פ",
            "H": "H",
            "I": "I",
            "J": "ſ",
            "K": "ʞ",
            "L": "˥",
            "M": "W",
            "N": "N",
            "O": "O",
            "P": "Ԁ",
            "Q": "Q",
            "R": "ᴿ",
            "S": "S",
            "T": "┴",
            "U": "∩",
            "V": "Λ",
            "W": "M",
            "X": "X",
            "Y": "⅄",
            "Z": "Z",
            "?": "¿",
            "!": "¡",
            ".": "˙",
            ",": "'",
        }

        # Separator characters
        self.separators = {
            UNICODE_FORMAT_STYLE.WIDE_SPACING: "​",  # Zero-width space
            UNICODE_FORMAT_STYLE.BLOCK_LETTERS: "█",  # Block character
            UNICODE_FORMAT_STYLE.INVISIBLE_SEPARATOR: "\u200d",  # Zero-width joiner
            UNICODE_FORMAT_STYLE.ZERO_WIDTH: "\u200e",  # Left-to-right mark
            UNICODE_FORMAT_STYLE.HAIR_SPACE: "\u200a",  # Hair space
            UNICODE_FORMAT_STYLE.THIN_SPACE: "\u2009",  # Thin space
        }

    def apply_combining_char(self, text: str, combining_char: str) -> str:
        """Apply combining character to each character in text"""
        return "".join(
            char + combining_char if char.isalnum() else char for char in text
        )

    def apply_character_map(
        self, text: str, char_map: dict, reverse: bool = False
    ) -> str:
        """Apply character mapping to text"""
        if reverse:
            text = text[::-1]
        result = "".join(char_map.get(char, char) for char in text)
        return result

    def apply_separator(self, text: str, separator: str) -> str:
        """Insert separator between each character"""
        if separator == "█":  # Special handling for block letters
            return separator.join(text)
        else:
            return separator.join(char for char in text if char != " ") if text else ""

    def format_text(self, text: str, style: UNICODE_FORMAT_STYLE) -> str:
        """Apply Unicode formatting style to text"""
        text = text.strip()
        if not text:
            return text

        if style == UNICODE_FORMAT_STYLE.STRIKETHROUGH:
            return self.apply_combining_char(text, self.combining_chars[style])
        elif style == UNICODE_FORMAT_STYLE.UNDERLINE:
            return self.apply_combining_char(text, self.combining_chars[style])
        elif style == UNICODE_FORMAT_STYLE.SUPERSCRIPT:
            return self.apply_character_map(text, self.superscript_map)
        elif style == UNICODE_FORMAT_STYLE.SUBSCRIPT:
            return self.apply_character_map(text, self.subscript_map)
        elif style == UNICODE_FORMAT_STYLE.UPSIDE_DOWN:
            return self.apply_character_map(text, self.upside_down_map, reverse=True)
        elif style in [
            UNICODE_FORMAT_STYLE.WIDE_SPACING,
            UNICODE_FORMAT_STYLE.BLOCK_LETTERS,
            UNICODE_FORMAT_STYLE.INVISIBLE_SEPARATOR,
            UNICODE_FORMAT_STYLE.ZERO_WIDTH,
            UNICODE_FORMAT_STYLE.HAIR_SPACE,
            UNICODE_FORMAT_STYLE.THIN_SPACE,
        ]:
            return self.apply_separator(text, self.separators[style])
        elif style == UNICODE_FORMAT_STYLE.MIXED_CASE:
            return "".join(
                char.upper() if i % 2 == 0 else char.lower()
                for i, char in enumerate(text)
                if char.isalpha()
            ) + "".join(char for char in text if not char.isalpha())
        return text


# =============================================================================
# TEXT DECORATION SYSTEM (for rich text formats)
# =============================================================================


class TEXT_DECORATION_STYLE(Enum):
    BOLD = "bold"
    ITALIC = "italic"
    UNDERLINE = "underline"
    STRIKETHROUGH = "strikethrough"
    SUPERSCRIPT = "superscript"
    SUBSCRIPT = "subscript"
    COLOR_RED = "color_red"
    COLOR_BLUE = "color_blue"
    COLOR_GREEN = "color_green"
    BACKGROUND_YELLOW = "bg_yellow"
    BACKGROUND_GRAY = "bg_gray"


class TextDecorationFormatter:
    def __init__(self, output_format: str = "html"):
        """
        Initialize formatter for specific output format

        Args:
            output_format: 'html', 'markdown', 'ansi', 'rich', 'excel'
        """
        self.output_format = output_format

        # Format templates for different outputs
        self.format_templates = {
            "html": {
                TEXT_DECORATION_STYLE.BOLD: ("<b>", "</b>"),
                TEXT_DECORATION_STYLE.ITALIC: ("<i>", "</i>"),
                TEXT_DECORATION_STYLE.UNDERLINE: ("<u>", "</u>"),
                TEXT_DECORATION_STYLE.STRIKETHROUGH: ("<s>", "</s>"),
                TEXT_DECORATION_STYLE.SUPERSCRIPT: ("<sup>", "</sup>"),
                TEXT_DECORATION_STYLE.SUBSCRIPT: ("<sub>", "</sub>"),
                TEXT_DECORATION_STYLE.COLOR_RED: (
                    '<span style="color: red;">',
                    "</span>",
                ),
                TEXT_DECORATION_STYLE.COLOR_BLUE: (
                    '<span style="color: blue;">',
                    "</span>",
                ),
                TEXT_DECORATION_STYLE.COLOR_GREEN: (
                    '<span style="color: green;">',
                    "</span>",
                ),
                TEXT_DECORATION_STYLE.BACKGROUND_YELLOW: (
                    '<span style="background: yellow;">',
                    "</span>",
                ),
                TEXT_DECORATION_STYLE.BACKGROUND_GRAY: (
                    '<span style="background: gray;">',
                    "</span>",
                ),
            },
            "markdown": {
                TEXT_DECORATION_STYLE.BOLD: ("**", "**"),
                TEXT_DECORATION_STYLE.ITALIC: ("*", "*"),
                TEXT_DECORATION_STYLE.UNDERLINE: ("<u>", "</u>"),  # HTML fallback
                TEXT_DECORATION_STYLE.STRIKETHROUGH: ("~~", "~~"),
                TEXT_DECORATION_STYLE.SUPERSCRIPT: ("<sup>", "</sup>"),
                TEXT_DECORATION_STYLE.SUBSCRIPT: ("<sub>", "</sub>"),
            },
            "ansi": {
                TEXT_DECORATION_STYLE.BOLD: ("\033[1m", "\033[0m"),
                TEXT_DECORATION_STYLE.ITALIC: ("\033[3m", "\033[0m"),
                TEXT_DECORATION_STYLE.UNDERLINE: ("\033[4m", "\033[0m"),
                TEXT_DECORATION_STYLE.STRIKETHROUGH: ("\033[9m", "\033[0m"),
                TEXT_DECORATION_STYLE.COLOR_RED: ("\033[91m", "\033[0m"),
                TEXT_DECORATION_STYLE.COLOR_BLUE: ("\033[94m", "\033[0m"),
                TEXT_DECORATION_STYLE.COLOR_GREEN: ("\033[92m", "\033[0m"),
                TEXT_DECORATION_STYLE.BACKGROUND_YELLOW: ("\033[103m", "\033[0m"),
                TEXT_DECORATION_STYLE.BACKGROUND_GRAY: ("\033[100m", "\033[0m"),
            },
            "rich": {
                TEXT_DECORATION_STYLE.BOLD: ("[bold]", "[/bold]"),
                TEXT_DECORATION_STYLE.ITALIC: ("[italic]", "[/italic]"),
                TEXT_DECORATION_STYLE.UNDERLINE: ("[underline]", "[/underline]"),
                TEXT_DECORATION_STYLE.STRIKETHROUGH: ("[strike]", "[/strike]"),
                TEXT_DECORATION_STYLE.COLOR_RED: ("[red]", "[/red]"),
                TEXT_DECORATION_STYLE.COLOR_BLUE: ("[blue]", "[/blue]"),
                TEXT_DECORATION_STYLE.COLOR_GREEN: ("[green]", "[/green]"),
            },
        }

    def format_text(self, text: str, style: TEXT_DECORATION_STYLE) -> str:
        """Apply text decoration style to text"""
        if self.output_format not in self.format_templates:
            return text

        style_map = self.format_templates[self.output_format]
        if style not in style_map:
            return text

        start_tag, end_tag = style_map[style]
        return f"{start_tag}{text}{end_tag}"

    def get_excel_formatting(self, style: TEXT_DECORATION_STYLE) -> Dict[str, Any]:
        """Get Excel formatting properties for openpyxl"""
        from openpyxl.styles import Font, PatternFill

        formatting = {}

        if style == TEXT_DECORATION_STYLE.BOLD:
            formatting["font"] = Font(bold=True)
        elif style == TEXT_DECORATION_STYLE.ITALIC:
            formatting["font"] = Font(italic=True)
        elif style == TEXT_DECORATION_STYLE.UNDERLINE:
            formatting["font"] = Font(underline="single")
        elif style == TEXT_DECORATION_STYLE.STRIKETHROUGH:
            formatting["font"] = Font(strike=True)
        elif style == TEXT_DECORATION_STYLE.COLOR_RED:
            formatting["font"] = Font(color="FF0000")
        elif style == TEXT_DECORATION_STYLE.COLOR_BLUE:
            formatting["font"] = Font(color="0000FF")
        elif style == TEXT_DECORATION_STYLE.COLOR_GREEN:
            formatting["font"] = Font(color="00FF00")
        elif style == TEXT_DECORATION_STYLE.BACKGROUND_YELLOW:
            formatting["fill"] = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
        elif style == TEXT_DECORATION_STYLE.BACKGROUND_GRAY:
            formatting["fill"] = PatternFill(
                start_color="808080", end_color="808080", fill_type="solid"
            )

        return formatting


# =============================================================================
# GENERATION FUNCTIONS
# =============================================================================


def generate_unicode_formatted_variations(
    question: str,
    language: Optional[LANGS] = LANGS.eng_Latn,
    styles: Optional[List[UNICODE_FORMAT_STYLE]] = None,
    n_variations: int = 1,
) -> List[str]:
    """Generate Unicode formatted variations"""

    if styles is None:
        styles = list(UNICODE_FORMAT_STYLE)

    formatter = UnicodeFormatter()
    variations = set()
    attempts = 0
    max_attempts = n_variations * 10

    while len(variations) < n_variations and attempts < max_attempts:
        style = random.choice(styles)
        formatted = formatter.format_text(question, style)
        variations.add(formatted)
        attempts += 1

    return list(variations)


def generate_text_decorated_variations(
    question: str,
    language: Optional[LANGS] = LANGS.eng_Latn,
    styles: Optional[List[TEXT_DECORATION_STYLE]] = None,
    output_format: str = "html",
    n_variations: int = 1,
) -> List[str]:
    """Generate text decoration formatted variations"""

    if styles is None:
        styles = list(TEXT_DECORATION_STYLE)

    formatter = TextDecorationFormatter(output_format)
    variations = set()
    attempts = 0
    max_attempts = n_variations * 10

    while len(variations) < n_variations and attempts < max_attempts:
        style = random.choice(styles)
        formatted = formatter.format_text(question, style)
        variations.add(formatted)
        attempts += 1

    return list(variations)


# =============================================================================
# EXCEL EXPORT FUNCTIONS
# =============================================================================


def write_unicode_to_excel(
    filename: str, texts: List[str], formatted_texts: List[List[str]]
):
    """Write Unicode formatted text to Excel (as plain text with Unicode characters)"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Unicode Formatted"

    # Headers
    ws["A1"] = "Original"
    ws["B1"] = "Unicode Formatted"

    for i, (original, variations) in enumerate(zip(texts, formatted_texts), 2):
        ws[f"A{i}"] = original
        ws[f"B{i}"] = variations[0] if variations else original

    wb.save(filename)


def write_decorated_to_excel(
    filename: str, texts: List[str], styles: List[TEXT_DECORATION_STYLE]
):
    """Write text with Excel formatting (actual rich text formatting)"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Text Decorated"

    formatter = TextDecorationFormatter()

    # Headers
    ws["A1"] = "Original"
    ws["B1"] = "Decorated"

    for i, (text, style) in enumerate(zip(texts, styles), 2):
        ws[f"A{i}"] = text
        ws[f"B{i}"] = text  # Set text first

        # Apply Excel formatting
        excel_formatting = formatter.get_excel_formatting(style)
        cell = ws[f"B{i}"]

        if "font" in excel_formatting:
            cell.font = excel_formatting["font"]
        if "fill" in excel_formatting:
            cell.fill = excel_formatting["fill"]

    wb.save(filename)


# =============================================================================
# EXAMPLE USAGE
# =============================================================================

if __name__ == "__main__":
    test_text = "Photosynthesis"

    print("=== UNICODE FORMATTING ===")
    unicode_variations = generate_unicode_formatted_variations(
        test_text,
        styles=[
            UNICODE_FORMAT_STYLE.STRIKETHROUGH,
            UNICODE_FORMAT_STYLE.SUPERSCRIPT,
            UNICODE_FORMAT_STYLE.UPSIDE_DOWN,
            UNICODE_FORMAT_STYLE.BLOCK_LETTERS,
        ],
        n_variations=4,
    )

    for i, variant in enumerate(unicode_variations, 1):
        print(f"{i}. {variant}")

    print("\n=== TEXT DECORATIONS (HTML) ===")
    html_variations = generate_text_decorated_variations(
        test_text,
        styles=[
            TEXT_DECORATION_STYLE.BOLD,
            TEXT_DECORATION_STYLE.ITALIC,
            TEXT_DECORATION_STYLE.COLOR_RED,
            TEXT_DECORATION_STYLE.STRIKETHROUGH,
        ],
        output_format="html",
        n_variations=4,
    )

    for i, variant in enumerate(html_variations, 1):
        print(f"{i}. {variant}")

    print("\n=== TEXT DECORATIONS (ANSI) ===")
    ansi_variations = generate_text_decorated_variations(
        test_text,
        styles=[
            TEXT_DECORATION_STYLE.BOLD,
            TEXT_DECORATION_STYLE.COLOR_RED,
            TEXT_DECORATION_STYLE.UNDERLINE,
        ],
        output_format="ansi",
        n_variations=3,
    )

    for i, variant in enumerate(ansi_variations, 1):
        print(f"{i}. {variant}")
