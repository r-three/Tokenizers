import hashlib
import json
import multiprocessing as mp
import unicodedata
from curses import meta
from dataclasses import dataclass, field
from functools import partial
from pathlib import Path
from tkinter import ALL
from typing import Any, Dict, List, Literal, Optional, Set

import numpy as np
import pandas as pd
import xlsxwriter
from click import Option
from regex import D

from xarch_tokenizers.config import Config
from xarch_tokenizers.experiment_config import load_config
from xarch_tokenizers.logging.logger import setup_logger
from xarch_tokenizers.perturbations import PERTURBATION_MAPPINGS
from xarch_tokenizers.perturbations.common import LANGS
from xarch_tokenizers.scripts.convert_dataset_to_hf_format import (
    cleanup_excel,
    read_data,
)


@dataclass
class PerturbationArgs(Config):
    dataset_path: Path = field(
        default=Path("input.jsonl"), metadata={"help": "Path to the input dataset."}
    )
    output_dir: Path = field(
        default=Path("output"), metadata={"help": "Output directory."}
    )
    output_format: Optional[Literal["jsonl", "xlsx"]] = field(
        default=None,
        metadata={"help": "Output format, leave empty to match input format.."},
    )

    perturbations: List[
        Literal[
            "cultural_references",
            "domain_specific_punctuation",
            "quotes_and_parentheses",
            "rare_n_grams",
            "sentence_boundaries",
            "wiki_jargon",
            "character_substitution",
            "deliberate_misspellings",
            "emoji_substitution",
            "word_reordering",
            "colloquial",
            "phonetic_spelling",
            "letter_repetition_for_emphasis",
            "word_concatenation",
            "borrowing",
            "calques",
            "transliteration",
            "chemical_formulas",
            "equations_with_mixed_notation",
            "numerical_formats",
            "scientific_notation",
            "unit_combinations",
            "clitics",
            "compounds",
            "contractions",
            "affixation_edge_cases",
            "derivations",
            "inflections",
            "academic_citations",
            "place_names_with_apostrophes",
            "special_names_across_cultures",
            "technical_product_names",
            "keyboard_proximity_errors",
            "noise_injection",
            "ocr_errors",
            "permutations",
            "typographical_errors",
            "abbreviations_with_periods",
            "brand_names_with_punctuation",
            "code_language_script_switching",
            "diacritics_presence_absence",
            "equivalent_expressions",
            "historical_spelling",
            "homoglyphs",
            "proper_nouns_with_unusual_capitalization",
            "regional_spelling_variations",
            "word_spacing_zero_width_characters_extra_space",
            "romanization",
            "capitalization",
            "email_addresses",
            "headers_and_section_titles",
            "list_markers",
            "text_decorations",
            "unicode_formatting",
            "unusual_formatting",
            "urls_and_file_paths",
        ]
    ] = field(default_factory=lambda: ["keyboard_proximity_errors"])

    sheet_name: str = field(
        default="English-Text-Completion",
        metadata={
            "help": "If the provided dataset_path is an excel file, pass the relevant sheet name."
        },
    )
    question_field: str = field(
        default="question",
        metadata={"help": "Name of the field containing the question."},
    )
    target_field: str = field(
        default="target",
        metadata={"help": "Name of the field containing the target."},
    )
    option_fields: List[str] = field(
        default_factory=lambda: ["A", "B", "C"],
        metadata={"help": "Names of the field containing the options."},
    )
    set_id_field: Optional[str] = field(
        default="Set Id", metadata={"help": "Name of the field containing the set id."}
    )
    variation_id_field: Optional[str] = field(
        default="Variation Id",
        metadata={"help": "Name of the field containing the variation id."},
    )
    perturbation_subcategory_field: Optional[str] = field(
        default="Subcategory",
        metadata={
            "help": "Name of the field containing the perturbation (subcategory)."
        },
    )
    perturbation_category_field: Optional[str] = field(
        default="Category",
        metadata={
            "help": "Name of the field containing the perturbation's general category."
        },
    )
    language_field: Optional[str] = field(
        default="Lang", metadata={"help": "Name of the field containing the language."}
    )

    def __post_init__(self):
        self.dataset_path = Path(self.dataset_path).resolve().absolute()
        self.output_dir = Path(self.output_dir).resolve().absolute()
        return super().__post_init__()


def read_data(config, logger):
    logger.info(f"Loading data from {config.dataset_path}")
    if config.dataset_path.suffix == ".xlsx":
        df = pd.read_excel(
            config.dataset_path,
            sheet_name=config.sheet_name,
            na_values=["", "null", "NULL"],
        )
        df = cleanup_excel(df, config)
    else:
        df = pd.read_csv(
            config.dataset_path,
            sep=config.separator,
            engine="python",
            na_values=["", "null", "NULL"],
        )

    logger.info(f"Loaded {len(df)} rows")

    # Create output directory
    output_dir_ = Path(config.output_dir)
    output_dir_.mkdir(parents=True, exist_ok=True)
    return df


def apply_perturbation(canonical_df, config, perturbation, perturbation_kwargs={}):
    if perturbation.func is None:
        return None
    perturbed_questions = canonical_df.apply(
        lambda row: perturbation.func(
            row[config.question_field],
            language=row.get(config.language_field, LANGS.eng_Latn),
            **perturbation_kwargs,
        ),
        axis=1,
    )
    if len(perturbed_questions) == 0:
        return None
    perturbed_df = canonical_df.copy()
    perturbed_df[config.question_field] = perturbed_questions
    perturbed_df[config.perturbation_subcategory_field] = perturbation.name
    perturbed_df[config.perturbation_category_field] = perturbation.category
    perturbed_df = perturbed_df.explode(config.question_field)
    perturbed_df = perturbed_df.reset_index(drop=True)

    # Remove rows where question is None/empty (if any variations failed)
    perturbed_df = perturbed_df.dropna(subset=[config.question_field])
    perturbed_df = perturbed_df[perturbed_df[config.question_field].str.strip() != ""]
    return perturbed_df


def save_with_xlsxwriter(df, output_path):
    """Save DataFrame with xlsxwriter which handles Unicode better"""

    # Use xlsxwriter engine which is more Unicode-friendly
    with pd.ExcelWriter(
        output_path,
        engine="xlsxwriter",  # options={"strings_to_unicode": True}
    ) as writer:
        df.to_excel(writer, index=False, sheet_name="Perturbed Data")

        # Optional: Set column widths for better display
        worksheet = writer.sheets["Perturbed Data"]
        for i, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
            worksheet.set_column(i, i, min(max_len, 50))  # Cap at 50 chars width


import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, PatternFill


def parse_ansi_codes(text):
    """Parse ANSI escape sequences and extract formatting info"""

    # ANSI code mappings
    ansi_to_format = {
        "0m": "reset",
        "1m": "bold",
        "3m": "italic",
        "4m": "underline",
        "9m": "strikethrough",
        "91m": "red",
        "92m": "green",
        "93m": "yellow",
        "94m": "blue",
        "95m": "magenta",
        "96m": "cyan",
        "100m": "bg_gray",
        "101m": "bg_red",
        "102m": "bg_green",
        "103m": "bg_yellow",
        "104m": "bg_blue",
    }

    # Find all ANSI sequences
    ansi_pattern = r"\[(\d+(?:;\d+)*)m"
    segments = []
    current_format = {}

    # Split text by ANSI codes
    parts = re.split(f"\033{ansi_pattern}", text)

    i = 0
    while i < len(parts):
        if i % 2 == 0:  # Text part
            if parts[i]:  # Not empty
                segments.append({"text": parts[i], "format": current_format.copy()})
        else:  # ANSI code part
            codes = parts[i].split(";")
            for code in codes:
                code_key = f"{code}m"
                if code_key in ansi_to_format:
                    format_type = ansi_to_format[code_key]
                    if format_type == "reset":
                        current_format = {}
                    else:
                        current_format[format_type] = True
        i += 1

    return segments


def create_rich_text_cell(segments):
    """Create Excel rich text from parsed segments"""
    if not segments:
        return ""

    # If only one segment with no formatting, return plain text
    if len(segments) == 1 and not segments[0]["format"]:
        return segments[0]["text"]

    rich_text = CellRichText()

    for segment in segments:
        text_part = segment["text"]
        format_info = segment["format"]

        if not format_info:
            # Plain text
            rich_text.append(text_part)
        else:
            # Create font with formatting
            font_kwargs = {}

            if "bold" in format_info:
                font_kwargs["bold"] = True
            if "italic" in format_info:
                font_kwargs["italic"] = True
            if "underline" in format_info:
                font_kwargs["underline"] = "single"
            if "strikethrough" in format_info:
                font_kwargs["strike"] = True

            # Colors
            color_map = {
                "red": "FF0000",
                "green": "008000",
                "blue": "0000FF",
                "yellow": "FFFF00",
                "magenta": "FF00FF",
                "cyan": "00FFFF",
            }

            for color, hex_code in color_map.items():
                if color in format_info:
                    font_kwargs["color"] = hex_code
                    break
            font = Font(**font_kwargs) if font_kwargs else Font()
            font = InlineFont(font)
            # if font_kwargs:
            #     print(font_kwargs, font)
            rich_text.append(TextBlock(font, text_part))

    return rich_text


def strip_ansi_codes(text):
    """Remove ANSI escape codes from text"""
    if not isinstance(text, str):
        return text
    ansi_escape = re.compile(r"\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])")
    return ansi_escape.sub("", text)


def save_with_ansi_formatting(df, output_path):
    """Save DataFrame converting ANSI codes to Excel formatting"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Formatted Data"

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=str(col_name))
        # Make headers bold
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    # Write data with formatting
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, str) and "\033[" in value:
                # Parse ANSI codes and create rich text
                segments = parse_ansi_codes(value)
                rich_text = create_rich_text_cell(segments)
                ws.cell(row=row_idx, column=col_idx, value=rich_text)
            else:
                # Plain text
                ws.cell(row=row_idx, column=col_idx, value=str(value))

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)


def perturb(config, logger):
    df = read_data(config, logger)
    logger.info("Data loaded successfully.")
    # Apply perturbations
    # get perturbation
    resulting_df = df.copy()
    canonical_ids = (
        df.groupby(config.set_id_field)[config.variation_id_field].min().reset_index()
    )
    canonical_df = df[
        df.apply(
            lambda x: x[config.variation_id_field]
            == canonical_ids.loc[
                canonical_ids[config.set_id_field] == x[config.set_id_field],
                config.variation_id_field,
            ].values[0],
            axis=1,
        )
    ]
    canonical_df[config.language_field] = canonical_df[config.language_field].map(
        lambda x: LANGS[x] if x in [l.value for l in LANGS] else LANGS.eng_Latn
    )
    for perturbation_name in config.perturbations:
        perturbation_kwargs = dict()
        if isinstance(perturbation_name, dict):
            perturbation_kwargs = list(perturbation_name.values())[0]
            perturbation_name = list(perturbation_name.keys())[0]
        if perturbation_name not in PERTURBATION_MAPPINGS:
            logger.warning(f"Unknown perturbation: {perturbation_name}")
            continue
        logger.info(f"Applying perturbation: {perturbation_name}")
        # Apply the specific perturbation to the dataframe
        perturbation = PERTURBATION_MAPPINGS[perturbation_name]
        if perturbation.func is None:
            logger.warning(
                f"Skipping non-automatable perturbation: {perturbation_name}"
            )
            continue
        perturbed_df = apply_perturbation(
            canonical_df, config, perturbation, perturbation_kwargs
        )
        if perturbed_df is None:
            continue
        resulting_df = pd.concat([resulting_df, perturbed_df], ignore_index=True)
    cnts = resulting_df.groupby(config.set_id_field)[config.variation_id_field].count()
    resulting_df = resulting_df.sort_values(config.set_id_field)
    resulting_df[config.variation_id_field] = [
        f"0.{i}" for set_id in cnts.index for i in range(cnts.loc[set_id])
    ]
    logger.info("All perturbations applied successfully.")

    # save output
    if (
        config.output_format is None and config.dataset_path.suffix == ".xlsx"
    ) or config.output_format == "xlsx":
        output_path = config.output_dir / "perturbed_data.xlsx"
        # Check if data contains ANSI codes
        has_ansi = (
            resulting_df.astype(str)
            .apply(lambda x: x.str.contains("\033\[", na=False))
            .any()
            .any()
        )

        if has_ansi:
            print("Converting ANSI formatting to Excel formatting...")
            save_with_ansi_formatting(resulting_df, output_path)
        else:
            print("No ANSI codes found.")
            with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
                resulting_df.to_excel(writer, index=False)

        # resulting_df.to_excel(output_path, index=False)
        # save_with_xlsxwriter(resulting_df, output_path)
    elif (
        config.output_format is None and config.dataset_path.suffix == ".jsonl"
    ) or config.output_format == "jsonl":
        output_path = config.output_dir / "perturbed_data.jsonl"
        resulting_df.to_json(output_path, orient="records", lines=True)
    logger.info(f"Perturbed data saved to {output_path}")
    return df


def main():
    config: PerturbationArgs = load_config(PerturbationArgs)
    logger = setup_logger(config, "perturber")

    # output_dir = config.output_dir.resolve().absolute()
    perturb(config, logger)


if __name__ == "__main__":
    main()
